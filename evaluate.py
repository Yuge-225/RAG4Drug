"""
evaluate.py
────────────────────────────────────────────────────────────────
LLM Alone vs RAG evaluation on eval_set.json.

For each question:
  1. LLM-alone response  (no context, raw GPT-4o-mini)
  2. RAG response        (SQL + vector retrieval + LLM)
  3. LLM-as-Judge scores both against DrugBank ground truth

Metrics (each 0–3 or 0/1):
  hallucination_score : fabrication level    (3=none, 0=severe)
  faithfulness_score  : alignment with GT    (3=full, 0=contradicts)
  completeness_score  : coverage of details  (3=comprehensive)
  severity_accuracy   : correct severity     (1/0, N/A for no_record)
  no_record_refusal   : correctly declined   (1/0, N/A for has_record)

Output:
  eval_results.json   raw scores + responses (checkpoint)
  eval_results.xlsx   3-sheet Excel report

Run:
  python evaluate.py
"""

import json
import time
import re
import sys
import platform
import psutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict

from dotenv import load_dotenv
load_dotenv()

from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage

EVAL_SET_PATH = Path("./eval_set.json")
RESULTS_PATH  = Path("./eval_results.json")
EXCEL_PATH    = Path("./eval_results.xlsx")
SLEEP_SEC     = 1.2   # seconds between API calls

_process = psutil.Process()


def system_info() -> dict:
    """Collect once at evaluation start and embed in results."""
    mem = psutil.virtual_memory()
    return {
        "python_version": sys.version.split()[0],
        "platform":       platform.platform(),
        "cpu_count":      psutil.cpu_count(logical=True),
        "total_ram_gb":   round(mem.total / 1024**3, 1),
        "model":          "gpt-4o-mini",
        "eval_date":      datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def measure_memory_mb() -> float:
    """Current process RSS memory in MB."""
    return round(_process.memory_info().rss / 1024**2, 1)


# ══════════════════════════════════════════════════════════════
# LLM Alone (no retrieval context)
# ══════════════════════════════════════════════════════════════

LLM_ALONE_SYSTEM = (
    "You are a clinical pharmacist AI assistant. "
    "Answer questions about drug interactions based on your training knowledge. "
    "Be specific about severity levels (Major/Moderate/Minor) when you know them. "
    "If you do not have reliable information about a specific drug combination, "
    "say so clearly rather than guessing."
)

def get_llm_alone_response(llm: ChatOpenAI, question: str) -> dict:
    """Returns dict with answer, latency_s, memory_mb."""
    messages = [
        SystemMessage(content=LLM_ALONE_SYSTEM),
        HumanMessage(content=question),
    ]
    mem_before = measure_memory_mb()
    t0 = time.perf_counter()
    content = llm.invoke(messages).content
    latency = round(time.perf_counter() - t0, 3)
    mem_after = measure_memory_mb()
    return {
        "answer":      content,
        "latency_s":   latency,
        "memory_mb":   mem_after,
        "memory_delta_mb": round(mem_after - mem_before, 1),
    }


# ══════════════════════════════════════════════════════════════
# RAG Response
# ══════════════════════════════════════════════════════════════

def get_rag_response(rag_service, question: str, session_id: str) -> dict:
    """Returns dict with answer, context_str, latency_s, memory_mb, hit."""
    mem_before = measure_memory_mb()
    t0 = time.perf_counter()
    context = rag_service.retrieve(question)
    response = rag_service.chain.invoke(
        {"question": question},
        config={"configurable": {"session_id": session_id}},
    )
    latency = round(time.perf_counter() - t0, 3)
    mem_after = measure_memory_mb()

    # Hit Rate: read from instance variable (raw dict, not parsed text)
    sql_result = getattr(rag_service, "last_sql_result", {})
    hit = len(sql_result.get("interactions", [])) > 0

    return {
        "answer":          response,
        "context":         context,
        "latency_s":       latency,
        "memory_mb":       mem_after,
        "memory_delta_mb": round(mem_after - mem_before, 1),
        "hit":             hit,
    }


# ══════════════════════════════════════════════════════════════
# LLM-as-Judge
# ══════════════════════════════════════════════════════════════

JUDGE_SYSTEM = (
    "You are an expert clinical pharmacology judge. "
    "Evaluate AI drug-interaction responses against DrugBank ground truth. "
    "Return ONLY valid JSON — no markdown, no extra text."
)

JUDGE_TEMPLATE = """\
QUESTION: {question}
CATEGORY: {category}

DRUGBANK GROUND TRUTH:
{gt_json}

AI RESPONSE TO EVALUATE:
{response}

Score this response. Return a JSON object with exactly these keys:

{{
  "hallucination_score": <integer 0-3>,
  "faithfulness_score":  <integer 0-3>,
  "completeness_score":  <integer 0-3>,
  "severity_accuracy":   <1 or 0 or null>,
  "no_record_refusal":   <1 or 0 or null>,
  "reasoning":           "<1-2 sentence justification>"
}}

Scoring rules:
hallucination_score:
  3 = no fabricated claims (everything stated is supported by GT or general pharmacology)
  2 = one or two minor unsupported details (e.g. mechanism not in GT but plausible)
  1 = notable unsupported claims that could mislead a clinician
  0 = significant fabrication (e.g. inventing an interaction that does not exist)

faithfulness_score:
  3 = response is fully consistent with DrugBank description
  2 = mostly consistent; minor omissions or slight paraphrasing
  1 = partially consistent; key details missing or altered
  0 = contradicts or completely ignores ground truth

completeness_score:
  3 = covers severity, mechanism, clinical advice
  2 = covers most key points; one area thin
  1 = covers some aspects; significant gaps
  0 = very superficial or off-topic

severity_accuracy:
  Set to 1 if the response correctly identifies the severity level matching GT (severe/moderate/minor).
  Set to 0 if severity is wrong or not mentioned when GT has a record.
  Set to null if category is "no_record".

no_record_refusal:
  Set to 1 if category is "no_record" AND response clearly states no interaction data found
    (phrases like "no record", "no data", "not found in database", "cannot confirm" count).
  Set to 0 if category is "no_record" AND response fabricates an interaction.
  Set to null for all other categories.

For multi_drug: evaluate against ALL pairs in the ground truth pairs list.\
"""

def _strip_markdown(text: str) -> str:
    text = text.strip()
    text = re.sub(r"^```[a-z]*\n?", "", text)
    text = re.sub(r"\n?```$", "", text)
    return text.strip()


def judge_response(judge_llm: ChatOpenAI, question: str, ground_truth: dict,
                   category: str, response: str) -> dict:
    gt_json = json.dumps(ground_truth, ensure_ascii=False, indent=2)
    prompt = JUDGE_TEMPLATE.format(
        question=question,
        category=category,
        gt_json=gt_json,
        response=response,
    )
    messages = [
        SystemMessage(content=JUDGE_SYSTEM),
        HumanMessage(content=prompt),
    ]
    raw = judge_llm.invoke(messages).content
    raw = _strip_markdown(raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {
            "hallucination_score": None,
            "faithfulness_score":  None,
            "completeness_score":  None,
            "severity_accuracy":   None,
            "no_record_refusal":   None,
            "reasoning": f"[JSON parse error] {raw[:300]}",
            "parse_error": True,
        }


# ══════════════════════════════════════════════════════════════
# Main evaluation loop
# ══════════════════════════════════════════════════════════════

def run_evaluation() -> list:
    from rag import RagService

    eval_set = json.loads(EVAL_SET_PATH.read_text(encoding="utf-8"))
    print(f"Loaded {len(eval_set)} evaluation items.\n")

    if RESULTS_PATH.exists():
        results = json.loads(RESULTS_PATH.read_text(encoding="utf-8"))
        done_ids = {r["id"] for r in results}
        print(f"Resuming: {len(done_ids)} items already done.\n")
    else:
        results  = []
        done_ids = set()

    llm       = ChatOpenAI(model="gpt-4o-mini", temperature=0)
    judge_llm = ChatOpenAI(model="gpt-4o-mini", temperature=0)

    print("Initialising RAG service...")
    rag = RagService()
    print("RAG ready.\n")

    sysinfo = system_info()
    print(f"System: {sysinfo['platform']}  Python {sysinfo['python_version']}  RAM {sysinfo['total_ram_gb']}GB\n")

    total = len(eval_set)
    for i, item in enumerate(eval_set, 1):
        item_id  = item["id"]
        question = item["question"]
        category = item["category"]
        gt       = item["ground_truth"]

        if item_id in done_ids:
            print(f"[{i:02d}/{total}] {item_id} — skip")
            continue

        sep = "─" * 70
        print(f"\n{sep}")
        print(f"[{i:02d}/{total}] {item_id}  [{category}]")
        print(f"  Q: {question[:80]}")
        print(sep)

        # Step 1: LLM alone
        print(f"  Step 1/4  LLM alone ...", end="", flush=True)
        try:
            llm_result = get_llm_alone_response(llm, question)
            print(f"  done  ({llm_result['latency_s']:.1f}s, {llm_result['memory_mb']:.0f}MB)")
        except Exception as e:
            llm_result = {"answer": f"ERROR: {e}", "latency_s": None,
                          "memory_mb": None, "memory_delta_mb": None}
            print(f"  ERROR: {e}")
        time.sleep(SLEEP_SEC)

        # Step 2: RAG retrieve + generate
        print(f"  Step 2/4  RAG retrieve + generate ...", end="", flush=True)
        try:
            rag_result = get_rag_response(rag, question, session_id=item_id)
            print(f"  done  ({rag_result['latency_s']:.1f}s, {rag_result['memory_mb']:.0f}MB,"
                  f" hit={rag_result['hit']})")
        except Exception as e:
            rag_result = {"answer": f"ERROR: {e}", "context": "", "latency_s": None,
                          "memory_mb": None, "memory_delta_mb": None, "hit": None}
            print(f"  ERROR: {e}")
        time.sleep(SLEEP_SEC)

        # Step 3: Judge LLM alone
        print(f"  Step 3/4  Judge LLM response ...", end="", flush=True)
        try:
            llm_scores = judge_response(judge_llm, question, gt, category, llm_result["answer"])
            print(f"  done  (hall={llm_scores.get('hallucination_score')},"
                  f" faith={llm_scores.get('faithfulness_score')},"
                  f" comp={llm_scores.get('completeness_score')})")
        except Exception as e:
            llm_scores = {"reasoning": f"judge error: {e}"}
            print(f"  ERROR: {e}")
        time.sleep(SLEEP_SEC)

        # Step 4: Judge RAG
        print(f"  Step 4/4  Judge RAG response ...", end="", flush=True)
        try:
            rag_scores = judge_response(judge_llm, question, gt, category, rag_result["answer"])
            print(f"  done  (hall={rag_scores.get('hallucination_score')},"
                  f" faith={rag_scores.get('faithfulness_score')},"
                  f" comp={rag_scores.get('completeness_score')})")
        except Exception as e:
            rag_scores = {"reasoning": f"judge error: {e}"}
            print(f"  ERROR: {e}")
        time.sleep(SLEEP_SEC)

        record = {
            "id":           item_id,
            "category":     category,
            "question":     question,
            "ground_truth": gt,
            "system_info":  sysinfo,
            "llm_response":        llm_result["answer"],
            "llm_latency_s":       llm_result["latency_s"],
            "llm_memory_mb":       llm_result["memory_mb"],
            "llm_memory_delta_mb": llm_result["memory_delta_mb"],
            "rag_response":        rag_result["answer"],
            "rag_context":         rag_result["context"][:3000],
            "rag_latency_s":       rag_result["latency_s"],
            "rag_memory_mb":       rag_result["memory_mb"],
            "rag_memory_delta_mb": rag_result["memory_delta_mb"],
            "rag_hit":             rag_result["hit"],
            "llm_scores":   llm_scores,
            "rag_scores":   rag_scores,
        }
        results.append(record)
        RESULTS_PATH.write_text(
            json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        print(f"  Checkpoint saved ({len(results)}/{total})")

    print(f"\nAll done. {len(results)} results saved to {RESULTS_PATH}")
    return results


# ══════════════════════════════════════════════════════════════
# Aggregation helpers
# ══════════════════════════════════════════════════════════════

def safe_mean(values: list) -> float | None:
    vals = [v for v in values if isinstance(v, (int, float))]
    return round(sum(vals) / len(vals), 4) if vals else None


def collect_metric(results: list, system: str, key: str,
                   category: str | None = None) -> list:
    subset = results if category is None else [r for r in results if r["category"] == category]
    return [r.get(system, {}).get(key) for r in subset]


# ══════════════════════════════════════════════════════════════
# Excel report
# ══════════════════════════════════════════════════════════════

def build_excel(results: list):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    LLM_FILL  = PatternFill("solid", fgColor="2E75B6")
    RAG_FILL  = PatternFill("solid", fgColor="70AD47")
    WHITE_FT  = Font(color="FFFFFF", bold=True)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_header_row(ws, row_num, fill):
        for cell in ws[row_num]:
            cell.fill  = fill
            cell.font  = WHITE_FT
            cell.alignment = CENTER

    CATEGORIES = ["severe", "moderate", "minor", "no_record", "multi_drug", "ALL"]
    METRICS = [
        # (display_name,        key,                 multiplier)
        ("Severity Accuracy (%)",     "severity_accuracy",   100),
        ("Hallucination Score (0-3)", "hallucination_score",   1),
        ("Faithfulness Score (0-3)",  "faithfulness_score",    1),
        ("Completeness Score (0-3)",  "completeness_score",    1),
        ("No-Record Refusal (%)",     "no_record_refusal",   100),
    ]

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.freeze_panes = "B3"

    row1 = [""]
    row2 = ["Metric"]
    for cat in CATEGORIES:
        row1 += [cat, ""]
        row2 += ["LLM", "RAG"]
    ws1.append(row1)
    ws1.append(row2)

    # merge category header cells
    col = 2
    for cat in CATEGORIES:
        ltr = get_column_letter(col)
        ws1.merge_cells(f"{ltr}1:{get_column_letter(col+1)}1")
        col += 2

    for mname, mkey, mult in METRICS:
        row = [mname]
        for cat in CATEGORIES:
            subset = results if cat == "ALL" else [r for r in results if r["category"] == cat]
            for sys_key in ("llm_scores", "rag_scores"):
                vals = [r.get(sys_key, {}).get(mkey) for r in subset]
                m = safe_mean(vals)
                row.append(round(m * mult, 2) if m is not None else "N/A")
        ws1.append(row)

    # Hit Rate row (RAG only)
    hr_row = ["Hit Rate (%, RAG only)"]
    for cat in CATEGORIES:
        subset = results if cat == "ALL" else [r for r in results if r["category"] == cat]
        has_rec_items = [r for r in subset
                         if r["ground_truth"].get("has_record") is True
                         or any(p["has_record"] for p in r["ground_truth"].get("pairs", []))]
        hr_row.append("N/A")
        hits = [r.get("rag_hit") for r in has_rec_items if r.get("rag_hit") is not None]
        m = safe_mean(hits)
        hr_row.append(round(m * 100, 2) if m is not None else "N/A")
    ws1.append(hr_row)

    # Latency row
    for label, key in [("Avg Latency (s)", "latency_s")]:
        row = [label]
        for cat in CATEGORIES:
            subset = results if cat == "ALL" else [r for r in results if r["category"] == cat]
            for prefix in ("llm", "rag"):
                vals = [r.get(f"{prefix}_{key}") for r in subset]
                m = safe_mean(vals)
                row.append(round(m, 2) if m is not None else "N/A")
        ws1.append(row)

    # Memory row
    for label, key in [("Avg Memory (MB)", "memory_mb")]:
        row = [label]
        for cat in CATEGORIES:
            subset = results if cat == "ALL" else [r for r in results if r["category"] == cat]
            for prefix in ("llm", "rag"):
                vals = [r.get(f"{prefix}_{key}") for r in subset]
                m = safe_mean(vals)
                row.append(round(m, 1) if m is not None else "N/A")
        ws1.append(row)

    style_header_row(ws1, 1, HDR_FILL)
    for j, cell in enumerate(ws1[2]):
        if j == 0:
            cell.fill = HDR_FILL
        elif j % 2 == 1:
            cell.fill = LLM_FILL
        else:
            cell.fill = RAG_FILL
        cell.font = WHITE_FT
        cell.alignment = CENTER

    ws1.column_dimensions["A"].width = 28
    for col_idx in range(2, 2 + len(CATEGORIES) * 2):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 10

    # Sample counts
    from collections import Counter
    ws1.append([])
    ws1.append(["Category", "N items"])
    cats_cnt = Counter(r["category"] for r in results)
    for cat in CATEGORIES[:-1]:
        ws1.append([cat, cats_cnt.get(cat, 0)])
    ws1.append(["TOTAL", len(results)])

    # System info block
    if results and results[0].get("system_info"):
        si = results[0]["system_info"]
        ws1.append([])
        ws1.append(["System Info", ""])
        for k, v in si.items():
            ws1.append([k, str(v)])

    # ── Sheet 2: Per-Question Results ─────────────────────────
    ws2 = wb.create_sheet("Per-Question Results")
    ws2.freeze_panes = "A2"

    hdrs = [
        "ID", "Category", "Question",
        "GT Severity", "GT Has Record", "RAG Hit",
        "LLM Latency(s)", "LLM Memory(MB)",
        "LLM Hallucination", "LLM Faithfulness", "LLM Completeness",
        "LLM Sev Acc", "LLM No-Rec Refusal", "LLM Reasoning",
        "RAG Latency(s)", "RAG Memory(MB)",
        "RAG Hallucination", "RAG Faithfulness", "RAG Completeness",
        "RAG Sev Acc", "RAG No-Rec Refusal", "RAG Reasoning",
    ]
    ws2.append(hdrs)
    style_header_row(ws2, 1, HDR_FILL)

    for r in results:
        gt = r["ground_truth"]
        if "pairs" in gt:
            sev_str = " | ".join(
                f"{p['drug_a']}×{p['drug_b']}:{p['severity']}"
                for p in gt["pairs"]
            )
            has_rec = any(p["has_record"] for p in gt["pairs"])
        else:
            sev_str = gt.get("severity")
            has_rec = gt.get("has_record")

        ls = r.get("llm_scores", {})
        rs = r.get("rag_scores", {})

        ws2.append([
            r["id"], r["category"], r["question"],
            sev_str, has_rec, r.get("rag_hit", "N/A"),
            r.get("llm_latency_s"), r.get("llm_memory_mb"),
            ls.get("hallucination_score"), ls.get("faithfulness_score"),
            ls.get("completeness_score"),  ls.get("severity_accuracy"),
            ls.get("no_record_refusal"),   ls.get("reasoning", ""),
            r.get("rag_latency_s"), r.get("rag_memory_mb"),
            rs.get("hallucination_score"), rs.get("faithfulness_score"),
            rs.get("completeness_score"),  rs.get("severity_accuracy"),
            rs.get("no_record_refusal"),   rs.get("reasoning", ""),
        ])

    widths2 = {"A": 10, "B": 12, "C": 38, "D": 22, "E": 12, "F": 9,
               "G": 13, "H": 14,
               "I": 14, "J": 14, "K": 14, "L": 12, "M": 15, "N": 38,
               "O": 13, "P": 14,
               "Q": 14, "R": 14, "S": 14, "T": 12, "U": 15, "V": 38}
    for col, w in widths2.items():
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: Raw Responses ────────────────────────────────
    ws3 = wb.create_sheet("Raw Responses")
    ws3.freeze_panes = "A2"

    ws3.append(["ID", "Category", "Question", "LLM Response", "RAG Response"])
    style_header_row(ws3, 1, HDR_FILL)

    for r in results:
        ws3.append([
            r["id"], r["category"], r["question"],
            r.get("llm_response", ""),
            r.get("rag_response", ""),
        ])
    for col, w in [("A", 10), ("B", 12), ("C", 40), ("D", 65), ("E", 65)]:
        ws3.column_dimensions[col].width = w

    wb.save(EXCEL_PATH)
    print(f"Excel report saved → {EXCEL_PATH}")


# ══════════════════════════════════════════════════════════════
# Terminal summary
# ══════════════════════════════════════════════════════════════

def print_summary(results: list):
    print("\n" + "=" * 55)
    print("EVALUATION SUMMARY  (higher = better for all metrics)")
    print("=" * 55)
    metric_labels = [
        ("Hallucination Score (0-3)", "hallucination_score"),
        ("Faithfulness Score  (0-3)", "faithfulness_score"),
        ("Completeness Score  (0-3)", "completeness_score"),
    ]
    print(f"{'Metric':<28} {'LLM':>8} {'RAG':>8}  {'Delta':>8}")
    print("-" * 55)
    for label, key in metric_labels:
        llm_m = safe_mean(collect_metric(results, "llm_scores", key))
        rag_m = safe_mean(collect_metric(results, "rag_scores", key))
        if llm_m is not None and rag_m is not None:
            delta = rag_m - llm_m
            print(f"{label:<28} {llm_m:>8.4f} {rag_m:>8.4f}  {delta:>+8.4f}")

    # Severity accuracy (has_record items only)
    has_rec = [r for r in results if r["ground_truth"].get("has_record") is True]
    if has_rec:
        llm_sa = safe_mean([r["llm_scores"].get("severity_accuracy") for r in has_rec])
        rag_sa = safe_mean([r["rag_scores"].get("severity_accuracy") for r in has_rec])
        if llm_sa is not None and rag_sa is not None:
            print(f"{'Severity Accuracy (%)':<28} {llm_sa*100:>7.1f}% {rag_sa*100:>7.1f}%  {(rag_sa-llm_sa)*100:>+7.1f}%")

    # No-record refusal (no_record items only)
    no_rec = [r for r in results if r["category"] == "no_record"]
    if no_rec:
        llm_nr = safe_mean([r["llm_scores"].get("no_record_refusal") for r in no_rec])
        rag_nr = safe_mean([r["rag_scores"].get("no_record_refusal") for r in no_rec])
        if llm_nr is not None and rag_nr is not None:
            print(f"{'No-Record Refusal (%)':<28} {llm_nr*100:>7.1f}% {rag_nr*100:>7.1f}%  {(rag_nr-llm_nr)*100:>+7.1f}%")

    # Latency & Memory
    llm_lat = safe_mean([r.get("llm_latency_s") for r in results])
    rag_lat = safe_mean([r.get("rag_latency_s") for r in results])
    if llm_lat and rag_lat:
        print(f"{'Avg Latency (s)':<28} {llm_lat:>8.2f} {rag_lat:>8.2f}  {rag_lat-llm_lat:>+8.2f}")

    llm_mem = safe_mean([r.get("llm_memory_mb") for r in results])
    rag_mem = safe_mean([r.get("rag_memory_mb") for r in results])
    if llm_mem and rag_mem:
        print(f"{'Avg Memory (MB)':<28} {llm_mem:>8.1f} {rag_mem:>8.1f}  {rag_mem-llm_mem:>+8.1f}")

    print("=" * 55)


# ══════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print(f"RAG Evaluation  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    results = run_evaluation()

    print("\nBuilding Excel report...")
    build_excel(results)

    print_summary(results)
    print("\nFiles written:")
    print(f"  {RESULTS_PATH}")
    print(f"  {EXCEL_PATH}")
